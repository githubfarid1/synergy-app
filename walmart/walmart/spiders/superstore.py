from datetime import datetime
import json
import time
import scrapy
from urllib.parse import urlencode, urlparse
from openpyxl import Workbook, load_workbook

class SuperstoreSpider(scrapy.Spider):
    name = "superstore"
    def __init__(self, name=None, **kwargs):
        super().__init__(name, **kwargs)

    def start_requests(self):
        # Creating URL for scrapings
        # scrapy crawl superstore -a xlsinput=C:/synergy-data-tester/Lookup Listing.xlsx" -a sheetname="Sheet1"
        # workbook = load_workbook(filename=self.xlsinput, read_only=False, keep_vba=True, data_only=True)
        workbook = load_workbook(filename=r"C:/synergy-data-tester/Lookup Listing.xlsx", read_only=False, keep_vba=True, data_only=True)
       
        # worksheet = workbook[self.sheetname]
        worksheet = workbook["Sheet1"]

        for i in range(2, worksheet.max_row + 1):
            url = worksheet[f'A{i}'].value
            domain = urlparse(url).netloc
            if domain == 'www.walmart.com' or domain == 'www.walmart.ca':
                yield scrapy.Request(
                    url=url,
                    callback=self.parse,
            )

    def parse(self, response):
        title = response.css("h1[data-automation=product-title] ::text").get()
        # title = response.css("h1::text").get()
        price = response.css("span[data-automation=buybox-price] ::text").get()
        # title = response.css("h1[itemprop=name] ::text").get()


        # script_tag = response.xpath('//script[@id="__NEXT_DATA__"]/text()').get()
        yield {
            "product_name":title,
            'price': price
        }
