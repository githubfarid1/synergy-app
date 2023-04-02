from openpyxl import Workbook, load_workbook

# filename1 = "/home/farid/dev/python/synergy-github/data/sample/copy_sheet/xUSA Small Shipment Creation V12.20.xlsm"
# filename2 = "/home/farid/dev/python/synergy-github/data/sample/copy_sheet/April 01 Labels.xlsx"
filename1 = r"C:/synergy-data-tester/copy_sheet/copy_sheet/USA Small Shipment Creation V12.20.xlsm"
filename2 = r"C:/synergy-data-tester/copy_sheet/copy_sheet/April 01 Labels.xlsx"


wb1 = load_workbook(filename=filename1, read_only=False, keep_vba=True, data_only=True)
wb2 = load_workbook(filename=filename2, read_only=False, keep_vba=True, data_only=True)
try:
    del wb1['New1']
except:
    pass

wb1.create_sheet('New1')
ws1 = wb1['New1']
ws2 = wb2["Sheet"]
for i in range(1, ws2.max_row + 1):
    cols = ('A', 'B', 'C', 'D', 'E', 'F', 'G', 'H')
    for col in cols:
        ws1['{}{}'.format(col, i)].value = ws2['{}{}'.format(col, i)].value
        print(ws2['{}{}'.format(col, i)].value)

wb1.save(filename1)
wb1.close()