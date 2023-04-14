from selenium import webdriver
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from webdriver_manager.chrome import ChromeDriverManager as CM
# from selenium.webdriver.support.select import Select
from selenium.webdriver.common.action_chains import ActionChains
import sys
import fitz
import os
import argparse
import time
from openpyxl import Workbook, load_workbook
# import unicodedata as ud
from sys import platform
import json
from random import randint
from datetime import date, datetime, timedelta
import warnings
import logging
from pathlib import Path
import amazon_lib as lib
import xlwings as xw
import shutil

logger = logging.getLogger()
logger.setLevel(logging.NOTSET)
logger2 = logging.getLogger()
logger2.setLevel(logging.NOTSET)


def clearlist(*args):
    for varlist in args:
        varlist.clear()

def explicit_wait():
    time.sleep(randint(1, 3))

def clear_screan():
    return
    try:
        if platform == "win32":
            os.system("cls")
        else:    
            os.system("clear")
    except Exception as er:
        print(er, "Command is not supported")

def pause(mess=""):
    input(mess)

def getConfig():
	file = open("setting.json", "r")
	config = json.load(file)
	return config

def getDownloadFolder():
    download_folder = os.path.expanduser('~/Downloads')    
    if platform == "win32":
        download_folder = os.getenv('USERPROFILE') + r'\Downloads'
    return download_folder

def killAllChrome():
    if platform == "win32":
        os.system("taskkill /f /im chrome.exe")

class AmazonShipment:
    def __init__(self, xlsfile, sname, chrome_data, download_folder, xlworkbook) -> None:
        try:
            xltmp = 'xlstmp' + xlsfile[-5:]
            self.__workbook = load_workbook(filename=xltmp, read_only=False, keep_vba=True, data_only=True)
            self.__worksheet = self.__workbook[sname]
            self.__xlworkbook = xlworkbook
            self.__xlworksheet = xlworkbook.sheets[sname]
        except Exception as e:
            logger.error(e)
            input("XLSX file or Sheet name not found")
            sys.exit()
        self.__datajson = json.loads("{}")
        self.__datalist = []
        self.__chrome_data = chrome_data
        # self.__download_folder = repr(download_folder)
        self.__download_folder = download_folder

        self.__xlsfile = xlsfile
        self.__delimeter = "/" 
        if platform == "win32":
            self.__delimeter = "\\"
        clear_screan()
        print("Kill All Chrome in the Background... ", end="")
        killAllChrome()
        print("passed")
 
        self.__driver = self.__browser_init()
        # input("pause")
        self.__data_generator()
        # exit()
        # self.__data_sanitizer()

    def __browser_init(self):
        config = getConfig()
        warnings.filterwarnings("ignore", category=UserWarning)
        options = webdriver.ChromeOptions()
        # options = Options()
        # options.add_argument("--headless")
        options.add_argument("user-data-dir={}".format(config['chrome_user_data'])) 
        options.add_argument("profile-directory={}".format(config['chrome_profile']))
        options.add_argument('--no-sandbox')
        options.add_argument("--log-level=3")
        # options.add_argument("--window-size=1200, 900")
        options.add_argument('--start-maximized')
        options.add_argument("--disable-notifications")
        options.add_experimental_option("excludeSwitches", ["enable-automation"])
        options.add_experimental_option('useAutomationExtension', False)
        profile = {"plugins.plugins_list": [{"enabled": False, "name": "Chrome PDF Viewer"}], # Disable Chrome's PDF Viewer
                    "download.default_directory": self.download_folder, # disable karena kadang gak jalan di PC lain. Jadi downloadnya tetap ke folder download default
                    "download.extensions_to_open": "applications/pdf",
                    "download.prompt_for_download": False,
                    'profile.default_content_setting_values.automatic_downloads': 1,
                    "download.directory_upgrade": True,
                    "plugins.always_open_pdf_externally": True #It will not show PDF directly in chrome                    
                    }
        options.add_experimental_option("prefs", profile)
        return webdriver.Chrome(service=Service(CM().install()), options=options)

    def parse(self):
        print("Try to login... ", end="")
        reslist = []
        '''
        # THIS METHOD WILL BE USE IF DIRECT ACCESS TO https://sellercentral.amazon.ca/fba/sendtoamazon?ref=fbacentral_nav_fba FAILED
        
        url = "https://sellercentral.amazon.ca/home"
        self.driver.get(url)
        try:
            WebDriverWait(self.driver, 10).until(EC.visibility_of_element_located((By.CSS_SELECTOR, "div[id='spacecasino-sellercentral-homepage-task-manager']")))
        except Exception as e:
            logger.error(e)
            print("Failed")
            input("Login Failed..")
            sys.exit()
        print("Passed")
        print("Go to Shipment Menu... ", end="")
        url = "https://sellercentral.amazon.ca/gp/ssof/shipping-queue.html/ref=xx_fbashipq_dnav_xx"
        
        self.driver.get(url)
        try:
            WebDriverWait(self.driver, 10).until(EC.visibility_of_element_located((By.CSS_SELECTOR, "div[id='tab-view']")))
        except Exception as e:
            logger.error(e)
            print("Failed")
            input("Shipment Menu Failed..")
            sys.exit()
        print("Passed")
        shadow_host = self.driver.find_element(By.CSS_SELECTOR, 'fba-navigation[active-tab="MANAGE_SHIPMENTS"')
        shadow_root = shadow_host.shadow_root
        shadow_content = shadow_root.find_element(By.CSS_SELECTOR, 'div[class="navigation"]')
        trial = 0
        while True:
            trial += 1
            try:
                a = ActionChains(self.driver)
                link = shadow_content.find_element(By.LINK_TEXT , 'Shipments')
                a.move_to_element(link).perform()
                explicit_wait()
                link = shadow_content.find_element(By.LINK_TEXT , 'Send to Amazon')
                link.click()
                break
            except:
                time.sleep(3)
                if trial >=5:
                    logger.error("Shipment menu Failed")
                    print("Failed")
                    input("Shipment Menu Failed..")
                    sys.exit()
                pass
        '''        
        
        url = "https://sellercentral.amazon.ca/fba/sendtoamazon?ref=fbacentral_nav_fba"
        self.driver.get(url)
        print("Check SKU page ready... ", end="")
        try:
            WebDriverWait(self.driver, 10).until(EC.visibility_of_element_located((By.CSS_SELECTOR, "div[data-testid='sku-list']")))
            checksku = self.driver.find_element(By.CSS_SELECTOR,"kat-tabs[id='skuTabs']").find_element(By.CSS_SELECTOR, "kat-tab-header[tab-id='3']").find_element(By.CSS_SELECTOR, "span[slot='label']").text
            if checksku != 'SKUs ready to send (0)':
                raise Exception('SKUs ready to send is not 0')
        except Exception as er:
            logger.error(er)
            logger.info("Trying to click start new link..")
            print("Trying to click start new link..", end="")
            try:
                WebDriverWait(self.driver, 10).until(EC.visibility_of_element_located((By.CSS_SELECTOR, "kat-link[data-testid='start-new-link']")))                
                self.driver.find_element(By.CSS_SELECTOR, "kat-link[data-testid='start-new-link']").click()
                print("Passed")
            except Exception as e:
                logger.error(e)
                print("Failed")
                sys.exit()
     
        print("")
        print("Starting Create Shipment...")
        # url = "https://sellercentral.amazon.com/fba/sendtoamazon?ref=fbacentral_nav_fba"
        # self.driver.get(url)
        try:
            WebDriverWait(self.driver, 10).until(EC.visibility_of_element_located((By.CSS_SELECTOR, "kat-link[data-testid='start-new-link']")))
        except:
            try:
                self.driver.find_element(By.CSS_SELECTOR, "input[id='signInSubmit']").click()
            except Exception as e:
                logger.error(e)
                input("Please click `Chrome Tester` menu, then login manually, then close the browser and try the script again")
                sys.exit()
        explicit_wait()
        self.driver.find_element(By.CSS_SELECTOR, "kat-link[data-testid='start-new-link']").click()
        WebDriverWait(self.driver, 10).until(EC.visibility_of_element_located((By.CSS_SELECTOR, "div[data-testid='sku-list']")))
        defsubmitter = self.driver.find_element(By.CSS_SELECTOR, "div[class='textBlock-60ch break-words']").text
        for idx, dlist in enumerate(self.datalist):
            # original_window = self.driver.current_window_handle
            submitter = dlist['submitter'].split("(")[0].strip()
            addresstmp = dlist['address']
            addresslist = addresstmp[addresstmp.find("(")+1:addresstmp.find(")")].strip().split(" ")
            address = addresslist[0] + " " + addresslist[1]# + " " + addresslist[2]
            # print(defsubmitter)
            print('#' * 5, dlist['name'], 'Start Process..', '#' * 5)
            if defsubmitter.find(submitter) != -1 and defsubmitter.find(address) != -1:
                # print('sama')
                print("Ship from label OK")
                pass
            else:
                print("Ship from Label Choosing..")
                explicit_wait()
                self.driver.find_element(By.CSS_SELECTOR, "a[data-testid='ship-from-another-address-link']").click()
                ck = WebDriverWait(self.driver, 20).until(EC.visibility_of_element_located((By.CSS_SELECTOR, "div[class='selected-address-tile']")))
                selects = self.driver.find_elements(By.CSS_SELECTOR, "div[class='address-tile']")
                for sel  in selects:
                    txt = sel.find_element(By.CSS_SELECTOR, "div[class='tile-address']").text
                    # print(txt)
                    if txt.find(submitter) != -1 and txt.find(address) != -1:
                        sel.find_element(By.CSS_SELECTOR, "button[class='secondary']").click()
                        # input('pause')
                        break
                explicit_wait()
                defsubmitter = self.driver.find_element(By.CSS_SELECTOR, "div[class='textBlock-60ch break-words']").text
                print("Ship from label OK")
            # PENTING
            # UNTUK CHROME DEVELOPER TOOLS AGAR BISA DEBUG SELECT
            # setTimeout(() => {debugger;}, 3000)                         
            explicit_wait()
            skuoption = WebDriverWait(self.driver, 20).until(EC.visibility_of_element_located((By.CSS_SELECTOR, "kat-dropdown[data-testid='search-dropdown']")))
            skuoption.click()
            explicit_wait()
            self.driver.find_element(By.CSS_SELECTOR, "kat-dropdown[data-testid='search-dropdown']").find_element(By.CSS_SELECTOR, "div[class='select-options']").find_element(By.CSS_SELECTOR, "div[class='option-inner-container']").find_element(By.CSS_SELECTOR, "div[data-value='MSKU']").click()
            # explicit_wait()
            for item in dlist['items']:
                skutxtsearch = WebDriverWait(self.driver, 20).until(EC.visibility_of_element_located((By.CSS_SELECTOR, "kat-input[data-testid='search-input']")))
                skutxtsearch.find_element(By.CSS_SELECTOR, "input").clear()
                xlssku = item['id'].upper()
                print('searching', xlssku, '..')
                skutxtsearch.find_element(By.CSS_SELECTOR, "input").send_keys(xlssku)
                explicit_wait()
                self.driver.find_element(By.CSS_SELECTOR, "a[data-testid='search-input-link']").click()
                WebDriverWait(self.driver, 20).until(EC.visibility_of_element_located((By.CSS_SELECTOR, "div[data-testid='sku-row-information-details']")))
                cols = self.driver.find_elements(By.CSS_SELECTOR, "div[data-testid='sku-row-information-details']")
                trial = 0
                while True:
                    trial += 1
                    try:
                        individual = WebDriverWait(cols[0], 20).until(EC.visibility_of_element_located((By.CSS_SELECTOR, "kat-dropdown[data-testid='packing-template-dropdown']")))
                        break
                    except:
                        time.sleep(3)
                        if trial >=5:
                            logger.error(xlssku + " Not found")
                            print(xlssku, "Not found")
                            input("Internet connection error, Script Failure..")
                            sys.exit()
                        pass
    
                if individual.text.find('Individual units') == -1:
                    individual.click()
                    explicit_wait()
                    wait = WebDriverWait(individual, 10).until(EC.visibility_of_element_located((By.CSS_SELECTOR, "div[class='select-options']")))
                    wait = WebDriverWait(wait, 10).until(EC.visibility_of_element_located((By.CSS_SELECTOR, "div[class='option-inner-container']")))
                    WebDriverWait(wait, 10).until(EC.visibility_of_element_located((By.CSS_SELECTOR, "div[data-name='Individual units']")))

                    individual.find_element(By.CSS_SELECTOR, "div[class='select-options']").find_element(By.CSS_SELECTOR, "div[class='option-inner-container']").find_element(By.CSS_SELECTOR, "div[data-name='Individual units']").click()
                explicit_wait()

                print(xlssku, "Input the unit number")
                numunit = cols[0].find_element(By.CSS_SELECTOR, "kat-input[data-testid='sku-readiness-number-of-units-input']").find_element(By.CSS_SELECTOR, "input[name='numOfUnits']")
                numunit.send_keys(item['total'])

                explicit_wait()
                try:
                    cols[0].find_element(By.CSS_SELECTOR, "kat-button[data-testid='skureadiness-confirm-button']").find_element(By.CSS_SELECTOR, "button[class='primary']").click()
                except:
                    pass

                explicit_wait()

                now = datetime.now()
                maxdate = now + timedelta(days=105)
                strexpiry = item['expiry'].strip()
                dformat = '%Y-%m-%d %H:%M:%S'
                dateinput = True
                if  strexpiry == 'None' or strexpiry == 'N/A':
                    dexpiry = now + timedelta(days=365)
                else:
                    try:
                        dexpiry = datetime.strptime(strexpiry, dformat)
                    except ValueError:
                        dateinput = False
                        error = True

                if dateinput == True:
                    if dexpiry < maxdate:
                        dexpiry = now + timedelta(days=365)
                        

                try:
                    expiry = dexpiry.strftime('%m/%d/%Y')
                except:
                    expiry = strexpiry

                try:
                    # expiry = "{}/{}/{}".format(item['expiry'][5:7], item['expiry'][8:10],item['expiry'][0:4])
                    inputexpiry = cols[0].find_element(By.CSS_SELECTOR, "kat-date-picker[id='expirationDatePicker']").find_element(By.CSS_SELECTOR, "input")
                    if inputexpiry.is_enabled():
                        print(xlssku, "Input the date expired")
                        inputexpiry.send_keys(expiry)
                        inputexpiry.send_keys(Keys.TAB)
                        explicit_wait()
                        wait = WebDriverWait(cols[0], 10).until(EC.visibility_of_element_located((By.CSS_SELECTOR, "kat-button[data-testid='skureadiness-confirm-button']")))
                        WebDriverWait(wait, 10).until(EC.visibility_of_element_located((By.CSS_SELECTOR, "button[class='primary']")))
                        cols[0].find_element(By.CSS_SELECTOR, "kat-button[data-testid='skureadiness-confirm-button']").find_element(By.CSS_SELECTOR, "button[class='primary']").click()
                except:
                    pass


            print(dlist['name'], 'Packaging..')
            # input("pause")
            time.sleep(2)
            self.driver.execute_script("window.scrollTo(0,document.body.scrollHeight)")
            wait = WebDriverWait(self.driver, 10).until(EC.visibility_of_element_located((By.XPATH , "//button[text()='Pack individual units']")))
            explicit_wait()
            wait.click()
            WebDriverWait(self.driver, 10).until(EC.visibility_of_element_located((By.CSS_SELECTOR, "div[data-testid='pack-group-controls']")))
            explicit_wait()
            # input("wait")
            print('Input box count, weight, dimension..')
            if dlist['boxcount'] == 1:
                WebDriverWait(self.driver, 10).until(EC.visibility_of_element_located((By.CSS_SELECTOR, "kat-label[text='Everything will fit into one box']")))
                self.driver.find_element(By.CSS_SELECTOR, "kat-label[text='Everything will fit into one box']").click()
                explicit_wait()
                WebDriverWait(self.driver, 10).until(EC.visibility_of_element_located((By.CSS_SELECTOR, "kat-button[data-testid='cli-input-method-verify-button']")))
                self.driver.find_element(By.CSS_SELECTOR, "kat-button[data-testid='cli-input-method-verify-button']").click()
                WebDriverWait(self.driver, 10).until(EC.visibility_of_element_located((By.CSS_SELECTOR, "div[data-testid='pack-group-cli-single-box-webform']")))
                weight = dlist['weightboxes'][0]
                dimension = dlist['dimensionboxes'][0].split("x")
                self.driver.find_element(By.CSS_SELECTOR, "kat-input[data-testid='cli-single-box-width-input']").find_element(By.CSS_SELECTOR, "input[type='number']").send_keys(dimension[0])
                self.driver.find_element(By.CSS_SELECTOR, "kat-input[data-testid='cli-single-box-height-input']").find_element(By.CSS_SELECTOR, "input[type='number']").send_keys(dimension[1])
                self.driver.find_element(By.CSS_SELECTOR, "kat-input[data-testid='cli-single-box-length-input']").find_element(By.CSS_SELECTOR, "input[type='number']").send_keys(dimension[2])
                explicit_wait()
                self.driver.find_element(By.CSS_SELECTOR, "kat-input[data-testid='cli-single-box-weight-input']").find_element(By.CSS_SELECTOR, "input[type='number']").send_keys(weight)
                explicit_wait()

                WebDriverWait(self.driver, 10).until(EC.visibility_of_element_located((By.CSS_SELECTOR, "kat-button[data-testid='cli-single-box-confirm-btn']")))
                self.driver.find_element(By.CSS_SELECTOR, "kat-button[data-testid='cli-single-box-confirm-btn']").click()
                explicit_wait()
                error = ''
                try:
                    error = self.driver.find_element(By.CSS_SELECTOR, "kat-alert[data-testid='pack-mixed-unit-error-results']").text
                except:
                    pass
                if error != '':
                    logger.error(error)
                    input(error)
                    sys.exit()
            else:
                WebDriverWait(self.driver, 10).until(EC.visibility_of_element_located((By.CSS_SELECTOR, "kat-label[text='Multiple boxes will be needed']")))

                self.driver.find_element(By.CSS_SELECTOR, "kat-label[text='Multiple boxes will be needed']").click()
                explicit_wait()

                WebDriverWait(self.driver, 10).until(EC.visibility_of_element_located((By.CSS_SELECTOR, "kat-button[data-testid='cli-input-method-verify-button']")))

                self.driver.find_element(By.CSS_SELECTOR, "kat-button[data-testid='cli-input-method-verify-button']").click()
                WebDriverWait(self.driver, 10).until(EC.visibility_of_element_located((By.CSS_SELECTOR, "kat-input[data-testid='cli-multi-box-webform-intial-container-quantity-input']")))
                explicit_wait()
                # input("pause")
                self.driver.find_element(By.CSS_SELECTOR, "kat-input[data-testid='cli-multi-box-webform-intial-container-quantity-input']").find_element(By.CSS_SELECTOR, "input[type='number']").send_keys(dlist['boxcount'])
                explicit_wait()
                wait = WebDriverWait(self.driver, 10).until(EC.visibility_of_element_located((By.CSS_SELECTOR, "kat-button[data-testid='cli-multi-box-open-webform-btn']")))
                WebDriverWait(wait, 10).until(EC.visibility_of_element_located((By.CSS_SELECTOR, "button[class='secondary']")))                
                self.driver.find_element(By.CSS_SELECTOR, "kat-button[data-testid='cli-multi-box-open-webform-btn']").find_element(By.CSS_SELECTOR, "button[class='secondary']").click()
                explicit_wait()
                cols = self.driver.find_element(By.CSS_SELECTOR, "div[data-testid='sku-quantity-inputs']").find_elements(By.CSS_SELECTOR, "div[class='flo-athens-border-bottom sku-input-child']")
                # print(cols)
                for col in cols:
                    explicit_wait()
                    tsku = col.find_element(By.CSS_SELECTOR, "div[data-testid='sku-information']").find_element(By.CSS_SELECTOR,"span[class='text-primary']").text.strip()
                    for item in dlist['items']:
                        txlssku = item['id'].strip().upper()
                        if tsku == txlssku:
                            cinputs = col.find_element(By.CSS_SELECTOR, "div[class='sku-quantity-wrapper']").find_elements(By.CSS_SELECTOR, "div[class='kat-input-padding-bottom-0 sku-input-katal-box']")
                            for idx, cinput in enumerate(cinputs):
                                cinput.find_element(By.CSS_SELECTOR, "input[type='number']").send_keys(item['boxes'][idx])


                for i in range(0,len(dlist['dimensionboxes'])-1 ):
                    wait = WebDriverWait(self.driver, 10).until(EC.visibility_of_element_located((By.CSS_SELECTOR, "div[class='bwd-add-dimension']")))
                    WebDriverWait(wait, 10).until(EC.visibility_of_element_located((By.CSS_SELECTOR, "kat-link[data-testid='bwd-add-dimension-link']")))
                    self.driver.find_element(By.CSS_SELECTOR, "div[class='bwd-add-dimension']").find_element(By.CSS_SELECTOR,"kat-link[data-testid='bwd-add-dimension-link']").click()

                cinputs = self.driver.find_element(By.CSS_SELECTOR, "div[data-testid='box-dimensions-labels']").find_elements(By.CSS_SELECTOR, "div[data-testid='box-dimensions-label']")
                for idx, cinput in enumerate(cinputs):
                    dimlist = dlist['dimensionboxes'][idx].split("x")
                    xinputs = cinput.find_elements(By.CSS_SELECTOR, "input[type='number']")
                    xinputs[0].send_keys(dimlist[0])
                    xinputs[1].send_keys(dimlist[1])
                    xinputs[2].send_keys(dimlist[2])
                    explicit_wait()

                bwdinput = self.driver.find_element(By.CSS_SELECTOR, "div[data-testid='bwd-input']") 
                cinputs = bwdinput.find_element(By.CSS_SELECTOR, "div[data-testid='box-weight-row']").find_elements(By.CSS_SELECTOR, "div[data-testid='weight-input-box']")
                for idx, cinput in enumerate(cinputs):
                    cinput.find_element(By.CSS_SELECTOR, "input[type='number']").send_keys(dlist['weightboxes'][idx])

                cinputs = bwdinput.find_elements(By.CSS_SELECTOR, "div[data-testid='bwd-input-child']")
                
                for idx, cinput in enumerate(cinputs):
                    explicit_wait()
                    xchecks = cinput.find_elements(By.CSS_SELECTOR, "kat-checkbox[data-testid='dimension-checkbox']")
                    xchecks[idx].click()
                    
                explicit_wait()
                waitme = WebDriverWait(self.driver, 10).until(EC.visibility_of_element_located((By.CSS_SELECTOR, "kat-modal-footer[data-testid='modal-footer']")))
                WebDriverWait(waitme, 20).until(EC.visibility_of_element_located((By.CSS_SELECTOR, "kat-button[data-testid='modal-confirm-button']")))

                self.driver.find_element(By.CSS_SELECTOR, "kat-modal-footer[data-testid='modal-footer']").find_element(By.CSS_SELECTOR, "kat-button[data-testid='modal-confirm-button']").click()

                try:
                    WebDriverWait(self.driver, 10).until(EC.visibility_of_element_located((By.CSS_SELECTOR, "kat-alert[data-testid='pack-group-cli-warning-results']")))
                    self.driver.find_element(By.CSS_SELECTOR, "kat-modal-footer[data-testid='modal-footer']").find_element(By.CSS_SELECTOR, "kat-button[data-testid='modal-confirm-button']").click()
                except:
                    pass

                WebDriverWait(self.driver, 10).until(EC.visibility_of_element_located((By.CSS_SELECTOR, "div[id='skudetails']")))
                explicit_wait()
            WebDriverWait(self.driver, 30).until(EC.element_to_be_clickable((By.CSS_SELECTOR, "kat-button[data-testid='confirm-and-continue']")))
            confirm = self.driver.find_element(By.CSS_SELECTOR, "kat-button[data-testid='confirm-and-continue']")
            confirm.click()
            WebDriverWait(self.driver, 120).until(EC.visibility_of_element_located((By.CSS_SELECTOR, "kat-button[data-testid='confirm-spd-shipping']")))

            print("input Send By Date, Shipping mode..")
            WebDriverWait(self.driver, 20).until(EC.visibility_of_element_located((By.CSS_SELECTOR, "div[data-testid='original-shipment']")))
            explicit_wait()
            todays_date = date.today()
            todays_str = "{}/{}/{}".format(str(todays_date.month), str(todays_date.day), str(todays_date.year))
            dateinput = self.driver.find_element(By.CSS_SELECTOR, "kat-date-picker[id='sendByDatePicker']").find_element(By.CSS_SELECTOR, "input")
            dateinput.clear()
            dateinput.send_keys(todays_str)
            explicit_wait()
            dateinput.send_keys(Keys.ESCAPE)
            # input("w")
            spd = WebDriverWait(self.driver, 30).until(EC.visibility_of_element_located((By.CSS_SELECTOR, "div[data-testid='shipping-mode-box-spd']")))
            spd.click()
            explicit_wait()
            self.driver.execute_script("window.scrollTo(0,document.body.scrollHeight)")
            print(dlist['name'], 'Saving the Shipping data')
            acc = WebDriverWait(self.driver, 120).until(EC.element_to_be_clickable((By.CSS_SELECTOR, "kat-button[data-testid='confirm-spd-shipping']")))
            acc.click()
            explicit_wait()
            print("Downloading PDF File to", self.download_folder)
            WebDriverWait(self.driver, 20).until(EC.visibility_of_element_located((By.CSS_SELECTOR, "div[data-testid='send-to-tile-list-row']")))
            self.driver.execute_script("window.scrollTo(0,document.body.scrollHeight)")
            WebDriverWait(self.driver, 600).until(EC.visibility_of_element_located((By.CSS_SELECTOR, "div[data-testid='print-section']")))            
            time.sleep(1)
            pdfs = self.driver.find_elements(By.CSS_SELECTOR, "div[data-testid='print-section']")
            original_window = self.driver.current_window_handle
            for pdf in pdfs:
                label = WebDriverWait(pdf, 20).until(EC.visibility_of_element_located((By.CSS_SELECTOR, "kat-dropdown[data-testid='print-label-dropdown']")))
                explicit_wait()
                label.click()
                explicit_wait()
                pdf.find_element(By.CSS_SELECTOR, "div[data-value='PackageLabel_Letter_2']").click()
                time.sleep(1)
                pdf.find_element(By.CSS_SELECTOR, "kat-button[data-testid='print-box-labels-button']").find_element(By.CSS_SELECTOR, "button[class='primary']").click()
                time.sleep(2)
                # self.driver.switch_to.window(self.driver.window_handles[1])
                # time.sleep(2)
                # self.driver.close()
                self.driver.switch_to.window(original_window)
                time.sleep(2)
            WebDriverWait(self.driver, 20).until(EC.visibility_of_element_located((By.CSS_SELECTOR, "kat-button[data-testid='proceed-tracking-details-button']")))
            explicit_wait()
            self.driver.find_element(By.CSS_SELECTOR, "kat-button[data-testid='proceed-tracking-details-button']").click()
            WebDriverWait(self.driver, 20).until(EC.visibility_of_element_located((By.CSS_SELECTOR, "div[data-testid='spd-tracking-table']")))
            explicit_wait()
            # print(dlist['name'], 'Saving to XLSX file..')
            tabs = self.driver.find_elements(By.CSS_SELECTOR, "div[data-testid='shipment-tracking-tab']")
            # tabcount = 0
            shiplist = []
            for tab in tabs:
                # tabcount += 1
                shipment_id = tab.find_elements(By.CSS_SELECTOR, "div")[3].text.replace("Shipment ID:","").strip()
                tab.click()
                explicit_wait()
                tracks = self.driver.find_element(By.CSS_SELECTOR, "div[data-testid='spd-tracking-table']").find_elements(By.CSS_SELECTOR,"kat-table-row[class='tracking-id-row']")
                dtmp = []
                for track in tracks:
                    trs = track.find_elements(By.CSS_SELECTOR, "kat-table-cell")
                    dict = {
                        'shipmentid': shipment_id,
                        'label':trs[1].text,
                        'trackid': trs[2].text,
                        'weight': trs[4].text,
                        'dimension': trs[5].text,

                    }
                    dtmp.append(dict)
                shiplist.append(dtmp)

            boxcols = ('E', 'F', 'G', 'H', 'I', 'J', 'K', 'L', 'M', 'N', 'O', 'P')
            stmp = []
            for ship in shiplist:
                for s in ship:
                    for boxcol in boxcols:
                        dimension = ""
                        weight = ""
                        box = str(self.worksheet['{}{}'.format(boxcol, dlist['begin'])].value)
                        # print(box)
                        if box != 'None':
                            dimrow = 0
                            for i in range(dlist['begin'], dlist['end']):
                                if self.worksheet['B{}'.format(i)].value == 'Weight':
                                    weight = self.worksheet['{}{}'.format(boxcol, i)].value
                                
                                if self.worksheet['B{}'.format(i)].value == 'Dimensions':
                                    dimension = self.worksheet['{}{}'.format(boxcol, i)].value
                                    dimrow = i
                                dimension = dimension.replace(" ","")
                                dimship = s['dimension'].replace(" ","")

                            if int(s['weight']) == int(weight) and dimension == dimship:
                                if not s['trackid'] in stmp and str(self.worksheet['{}{}'.format(boxcol, dimrow+2)].value) == 'None':
                                    stmp.append(s['trackid'])
                                    self.__extract_pdf(box=box, shipment_id=s['shipmentid'], label=s['label'])
                                    # self.worksheet['{}{}'.format(boxcol, dimrow+1)].value = s['label']
                                    # self.worksheet['{}{}'.format(boxcol, dimrow+2)].value = s['trackid']
                                    # restup = (f"{boxcol}{dimrow+1}", s['label'], f"{boxcol}{dimrow+2}", s['trackid'])
                                    # reslist.append(restup)
                                    self.xlworksheet[f"{boxcol}{dimrow+1}"].value = s['label']
                                    self.xlworksheet[f"{boxcol}{dimrow+2}"].value = s['trackid']
            # self.workbook.save(self.xlsfile)
            # print(dlist['name'], 'Saved to', self.xlsfile)
            print(dlist['name'], 'Extract PDF..')

            print('#' * 5, dlist['name'], "End Process", '#' * 5)
            logger2.info(dlist['name'] + " Created..")
            explicit_wait()
            print("Processing next shipping..", end="\n\n")
            explicit_wait()
            WebDriverWait(self.driver, 10).until(EC.visibility_of_element_located((By.CSS_SELECTOR, "kat-link[data-testid='start-new-link']")))
            explicit_wait()
            self.driver.find_element(By.CSS_SELECTOR, "kat-link[data-testid='start-new-link']").click()
            
            WebDriverWait(self.driver, 10).until(EC.visibility_of_element_located((By.CSS_SELECTOR, "div[data-testid='sku-list']")))
            explicit_wait()
            # close all download windows 
            original_window = self.driver.current_window_handle
            for handle in self.driver.window_handles:
                if handle != original_window:
                    self.driver.switch_to.window(handle)
                    self.driver.close()
            self.driver.switch_to.window(original_window)
        # self.workbook.save(self.xlsfile)
        # self.xlworkbook.save(self.xlsfile)
        # self.xlworkbook.close()
        # self.workbook.close()
        print('Saved All Shipment to', self.xlsfile)
        self.driver.quit()
        print('All Shipment has Created...')
        
    def __extract_pdf(self, box, shipment_id, label):
        pdffile = "{}{}package-{}.pdf".format(self.download_folder, self.file_delimeter, shipment_id)
        foldername = "{}{}combined".format(self.download_folder, self.file_delimeter) 
        isExist = os.path.exists(foldername)
        if not isExist:
            os.makedirs(foldername)
        white = fitz.utils.getColor("white")
        mfile = fitz.open(pdffile)
        fname = "{}{}{}.pdf".format(foldername, self.file_delimeter,  box.strip())
        tmpname = "{}{}{}.pdf".format(foldername, self.file_delimeter, "tmp")

        found = False
        pfound = 0
        for i in range(0, mfile.page_count):
            page = mfile[i]
            plist = page.search_for(label)
            if len(plist) != 0:
                found = True
                pfound = i
                break
        if found:
            single = fitz.open()
            single.insert_pdf(mfile, from_page=pfound, to_page=pfound)
            mfile.close()
            single.save(tmpname)
            mfile = fitz.open(tmpname)
            page = mfile[0]
            page.insert_text((550.2469787597656, 100.38037109375), "Box:{}".format(str(box)), rotate=90, color=white)
            page.set_rotation(90)
            mfile.save(fname)

    def __data_generator(self):
        print("Data Mounting... ", end="")
        shipmentlist = []
        for i in range(2, self.worksheet.max_row + 1):
            shipment_row = str(self.worksheet['A{}'.format(i)].value)
            if shipment_row.find('Shipment') != -1:
                # print(shipment_row, i)
                startrow = i
                y = i
                shipment_empty = True
                while True:
                    y += 1

                    # skip if shipment_id was filled
                    if ''.join(str(self.worksheet['B{}'.format(y)].value)).strip() == 'Shipment ID':
                        if ''.join(str(self.worksheet['E{}'.format(y)].value)).strip() != 'None':
                            shipment_empty = False

                    if str(self.worksheet['B{}'.format(y)].value) == 'Tracking Number':
                        endrow = y + 1
                        i = y + 1
                        break
                if shipment_empty == True:
                    shipmentlist.append({'begin':startrow, 'end':endrow})
                else:
                    logger2.info(shipment_row + " Skipped")

        # print(json.dumps(shipmentlist))
        for index, shipmentdata in enumerate(shipmentlist):
            shipmentlist[index]['submitter'] = self.worksheet['B{}'.format(shipmentdata['begin'])].value
            shipmentlist[index]['address'] = self.worksheet['B{}'.format(shipmentdata['begin']+1)].value
            shipmentlist[index]['name'] = self.worksheet['A{}'.format(shipmentdata['begin'])].value
            boxes = ('E', 'F', 'G', 'H', 'I', 'J', 'K', 'L', 'M', 'N', 'O', 'P')
            boxcount = 0
            for box in boxes:
                
                if self.worksheet['{}{}'.format(box, shipmentdata['begin'])].value != None:
                    boxcount += 1
                else:
                    break
            if boxcount == 0:
                del shipmentlist[index]
                continue
            shipmentlist[index]['boxcount'] = boxcount
            start = shipmentdata['begin'] + 2
            shipmentlist[index]['weightboxes'] = []
            shipmentlist[index]['dimensionboxes'] = []
            shipmentlist[index]['nameboxes'] = []
            shipmentlist[index]['items'] = []

            # get weightboxes
            rowsearch = 0
            for i in range(start, shipmentdata['end']):
                if self.worksheet['B{}'.format(i)].value == 'Weight':
                    rowsearch = i
                    break
            
            for ke, box in enumerate(boxes):
                if ke == boxcount:
                    break
                shipmentlist[index]['weightboxes'].append(self.worksheet['{}{}'.format(box, rowsearch)].value)

            # get dimensionboxes
            rowsearch = 0
            for i in range(start, shipmentdata['end']):
                if self.worksheet['B{}'.format(i)].value == 'Dimensions':
                    rowsearch = i
                    break
            
            for ke, box in enumerate(boxes):
                if ke == boxcount:
                    break
                shipmentlist[index]['dimensionboxes'].append(self.worksheet['{}{}'.format(box, rowsearch)].value)

            #get nameboxes
            for ke, box in enumerate(boxes):
                if ke == boxcount:
                    break
                shipmentlist[index]['nameboxes'].append(str(self.worksheet['{}{}'.format(box, shipmentdata['begin'])].value))

            ti = -1
            for i in range(start, shipmentdata['end']):
                ti += 1
                if self.worksheet['A{}'.format(i)].value == None or str(self.worksheet['A{}'.format(i)].value).strip() == '':
                    break
                # shipmentlist[index]['items'].append()
                
                dict = {
                    'id': self.worksheet['A{}'.format(i)].value,
                    'name': self.worksheet['B{}'.format(i)].value,
                    'total': self.worksheet['C{}'.format(i)].value,
                    'expiry': str(self.worksheet['D{}'.format(i)].value),
                    'boxes':[],

                }

                shipmentlist[index]['items'].append(dict)
                for ke, box in enumerate(boxes):
                    if ke == boxcount:
                        break
                    if self.worksheet['{}{}'.format(box, i)].value == None or str(self.worksheet['{}{}'.format(box, i)].value).strip() == '':
                        shipmentlist[index]['items'][ti]['boxes'].append(0)
                    else:                           
                        shipmentlist[index]['items'][ti]['boxes'].append(self.worksheet['{}{}'.format(box, i)].value)

        
        #cleansing
        idxdel = []
        for index, shipmentdata in enumerate(shipmentlist):
            try:
                cheat = shipmentdata['name']
            except:
                idxdel.append(index)
        
        for idx in idxdel:
            for index, shipmentdata in enumerate(shipmentlist):
                try:
                    cheat = shipmentdata['name']
                except:
                    del shipmentlist[index]
                
            # pass
        
        self.datalist = shipmentlist
        explicit_wait()
        print("Passed")

    def data_sanitizer(self):
        print("Try to login... ", end="")
        url = "https://sellercentral.amazon.ca/fba/sendtoamazon?ref=fbacentral_nav_fba"
        self.driver.get(url)
        print("Check SKU page ready... ", end="")
        try:
            WebDriverWait(self.driver, 10).until(EC.visibility_of_element_located((By.CSS_SELECTOR, "div[data-testid='sku-list']")))
            checksku = self.driver.find_element(By.CSS_SELECTOR,"kat-tabs[id='skuTabs']").find_element(By.CSS_SELECTOR, "kat-tab-header[tab-id='3']").find_element(By.CSS_SELECTOR, "span[slot='label']").text
            if checksku != 'SKUs ready to send (0)':
                raise Exception('SKUs ready to send is not 0')
        except Exception as er:
            logger.error(er)
            logger.info("Trying to click start new link..")
            print("Trying to click start new link..", end="")
            try:
                self.driver.find_element(By.CSS_SELECTOR, "kat-link[data-testid='start-new-link']").click()
                print("Passed")
            except Exception as e:
                logger.error(e)
                print("Failed")
                sys.exit()

        explicit_wait()
        print('Checking Excel Data..')
        self.driver.find_element(By.CSS_SELECTOR, "kat-dropdown[data-testid='search-dropdown']").click()
        explicit_wait()
        self.driver.find_element(By.CSS_SELECTOR, "kat-dropdown[data-testid='search-dropdown']").find_element(By.CSS_SELECTOR, "div[class='select-options']").find_element(By.CSS_SELECTOR, "div[class='option-inner-container']").find_element(By.CSS_SELECTOR, "div[data-value='MSKU']").click()
        explicit_wait()
        idxdel = []

        for idx, dlist in enumerate(self.datalist):
            print(dlist['name'], "... ", end="")
            shipping_name = dlist['name']
            error = False
            errorlist = []
            notelist = []
            defsubmitter = self.driver.find_element(By.CSS_SELECTOR, "div[class='textBlock-60ch break-words']").text

            submitter = dlist['submitter'].split("(")[0].strip()
            addresstmp = dlist['address']
            addresslist = addresstmp[addresstmp.find("(")+1:addresstmp.find(")")].strip().split(" ")
            address = addresslist[0] + " " + addresslist[1]# + " " + addresslist[2]
            address_found = False
            if defsubmitter.find(submitter) != -1 and defsubmitter.find(address) != -1:
                address_found = True
            else:
                addresslink = WebDriverWait(self.driver, 10).until(EC.visibility_of_element_located((By.CSS_SELECTOR, "a[data-testid='ship-from-another-address-link']")))
                addresslink.click()
                try:
                    WebDriverWait(self.driver, 30).until(EC.visibility_of_element_located((By.CSS_SELECTOR, "div[class='selected-address-tile']")))
                except:
                    input("Address list failed to open")
                    sys.exit()

                selects = self.driver.find_elements(By.CSS_SELECTOR, "div[class='address-tile']")
                explicit_wait()
                address_found = False

                for sel  in selects:
                    txt = sel.find_element(By.CSS_SELECTOR, "div[class='tile-address']").text
                    if txt.find(submitter) != -1 and txt.find(address) != -1:
                        sel.find_element(By.CSS_SELECTOR, "button[class='secondary']").click()
                        address_found = True
                        break

            if not address_found:
                errorlist.append("Address or Submitter not Found")
                error = True
                self.driver.find_element(By.CSS_SELECTOR, "div[class='selected-address-tile']").find_element(By.CSS_SELECTOR, "button[class='secondary']").send_keys(Keys.ESCAPE)
            
            if len(dlist['dimensionboxes']) == 0:
                error = True
                errmsg = "dimension value is Empty"
                errorlist.append(errmsg)

            for dim in dlist['dimensionboxes']:
                if lib.checkdimension(dim) == False:
                    error = True
                    errmsg = "{} dimension box value is wrong".format(dim)
                    errorlist.append(errmsg)
 
            wbox = "".join(str(x) for x in dlist['weightboxes'])
            if wbox.isnumeric() == False:
                error = True
                errmsg = "Weight box value is wrong"
                errorlist.append(errmsg)

            for idx2, item in enumerate(dlist['items']):
                self.driver.find_element(By.CSS_SELECTOR,"kat-input[data-testid='search-input']").find_element(By.CSS_SELECTOR, "input").clear()
                xlssku = item['id'].upper()
                self.driver.find_element(By.CSS_SELECTOR,"kat-input[data-testid='search-input']").find_element(By.CSS_SELECTOR, "input").send_keys(xlssku)
                # explicit_wait()
                searchinput = WebDriverWait(self.driver, 20).until(EC.element_to_be_clickable((By.CSS_SELECTOR, "a[data-testid='search-input-link']")))
                searchinput.click()
                explicit_wait()
                cols = self.driver.find_elements(By.CSS_SELECTOR, "div[data-testid='sku-row-information-details']")
                sku = ''
                try:
                    sku = cols[0].find_element(By.CSS_SELECTOR, "div[data-testid='msku']").find_element(By.CSS_SELECTOR, "span").text
                except:
                    error = True
                    errorlist.append(xlssku + ' Not Found')

                if xlssku != sku:
                    errorlist.append(sku + ' Not Match')
                    error = True
                errormsg = ''
                try:
                    errormsg = cols[0].find_element(By.CSS_SELECTOR, "div[data-testid='sku-action-info']").find_element(By.CSS_SELECTOR, "span[data-testid='sku-action-error-text']").text
                except:
                    pass
                if errormsg != '':
                    errorlist.append(errormsg)
                    error = True

                try:
                    individual = WebDriverWait(cols[0], 20).until(EC.visibility_of_element_located((By.CSS_SELECTOR, "kat-dropdown[data-testid='packing-template-dropdown']")))
                except Exception as e:
                    logger.error(e)
                    error = True
                    # print(xlssku, " Not Found")
                    idxdel.append(dlist['name'])
                    continue
                    # sys.exit()

                if individual.text.find('Individual units') == -1:
                    individual.click()
                    explicit_wait()
                    individual.find_element(By.CSS_SELECTOR, "div[class='select-options']").find_element(By.CSS_SELECTOR, "div[class='option-inner-container']").find_element(By.CSS_SELECTOR, "div[data-name='Individual units']").click()

                if cols[0].find_element(By.CSS_SELECTOR, "div[data-testid='sku-action-info'").text.find('Prep not required') == -1:
                    try:
                        try:
                            infoprep = cols[0].find_element(By.CSS_SELECTOR, "kat-link[data-testid='sku-action-info-prep-missing-link']")
                        except:
                            infoprep = cols[0].find_element(By.CSS_SELECTOR, "kat-link[data-testid='prep-modal-link']")
                        explicit_wait()
                        infoprep.click()
                        catprep = self.driver.find_element(By.CSS_SELECTOR, "kat-dropdown[data-testid='prep-guidance-prep-category-dropdown']")
                        explicit_wait()
                        catprep.click()
                        catprep.find_element(By.CSS_SELECTOR, "div[class='select-options']").find_element(By.CSS_SELECTOR, "div[class='option-inner-container']").find_element(By.CSS_SELECTOR, "div[data-value='NONE']").click()
                        explicit_wait()
                        self.driver.find_element(By.CSS_SELECTOR, "kat-modal-footer[data-testid='modal-footer']").find_element(By.CSS_SELECTOR, "kat-button[variant='primary']").find_element(By.CSS_SELECTOR, "button[class='primary']").click()
                        explicit_wait()
                        self.driver.find_element(By.CSS_SELECTOR, "kat-modal-footer[data-testid='modal-footer']").find_element(By.CSS_SELECTOR, "kat-button[data-testid='packing-template-save-button']").find_element(By.CSS_SELECTOR, "button[class='primary']").click()
                    except:
                        try:
                            self.driver.find_element(By.CSS_SELECTOR, "kat-button[data-testid='prep-category-update-btn']").find_element(By.CSS_SELECTOR, "button[class='primary']").click()
                            explicit_wait()
                            self.driver.find_element(By.CSS_SELECTOR, "kat-button[data-testid='packing-template-save-button']").find_element(By.CSS_SELECTOR, "button[class='primary']").click()
                        except:                        
                            pass
                try:
                    WebDriverWait(cols[0], 10).until(EC.text_to_be_present_in_element((By.CSS_SELECTOR , "span[data-testid='prep-fee-text']"), "Prep not required"))
                except:    
                    try:
                        WebDriverWait(cols[0], 10).until(EC.text_to_be_present_in_element((By.CSS_SELECTOR , "span[data-testid='prep-fee-text']"), "Unit prep: By seller"))
                    except:
                        try:
                            self.driver.find_element(By.CSS_SELECTOR, "kat-button[data-testid='packing-template-cancel-button']").find_element(By.CSS_SELECTOR, "button[class='secondary']").click()
                            error = True
                            idxdel.append(dlist['name'])
                        except:
                            error = True
                            idxdel.append(dlist['name'])
                            pass


                # except:
                #     try:
                #         WebDriverWait(cols[0], 10).until(EC.text_to_be_present_in_element((By.CSS_SELECTOR , "span[data-testid='prep-fee-text']"), "Unit prep: By seller"))
                #     except:
                #         pass

                try:
                    numunit = cols[0].find_element(By.CSS_SELECTOR, "kat-input[data-testid='sku-readiness-number-of-units-input']").find_element(By.CSS_SELECTOR, "input[name='numOfUnits']")
                    if not numunit.is_enabled():
                        errorlist.append(sku + " unit number disabled")
                        error = True                     
                except:
                    errorlist.append(sku + " unit number disabled")
                    error = True

                now = datetime.now()
                maxdate = now + timedelta(days=105)
                strexpiry = item['expiry'].strip()
                dformat = '%Y-%m-%d %H:%M:%S'
                dateinput = True
                if  strexpiry == 'None' or strexpiry == 'N/A':
                    dexpiry = now + timedelta(days=365)
                    notelist.append('{}: date expiry is empty, it will be adjusted to now +1 year'.format(xlssku))

                else:
                    try:
                        dexpiry = datetime.strptime(strexpiry, dformat)
                    except ValueError:
                        dateinput = False
                        error = True
                        errorlist.append("{}: wrong date expiry value".format(xlssku))                

                if dateinput == True:
                    if dexpiry < maxdate:
                        dexpiry = now + timedelta(days=365)
                        notelist.append("{}: date expiry is less than 105 days, it will be adjusted to now +1 year".format(xlssku))

                try:
                    expiry = dexpiry.strftime('%m/%d/%Y')
                except:
                    expiry = strexpiry

                try:
                    inputexpiry = cols[0].find_element(By.CSS_SELECTOR, "kat-date-picker[id='expirationDatePicker']").find_element(By.CSS_SELECTOR, "input")
                    if inputexpiry.is_enabled():
                        inputexpiry.send_keys(expiry)
                        inputexpiry.send_keys(Keys.TAB)
                except:
                    pass

                errormsg = ''
                try:
                    errormsg = cols[0].find_element(By.CSS_SELECTOR, "kat-label[data-testid='sku-readiness-expiration-date-error']").text
                except:
                    pass
                if errormsg != '':
                    errorlist.append(errormsg)
                    error = True
                
                boxes = "".join(str(x) for x in item['boxes'])
                if boxes.isnumeric() == False:
                    error = True
                    errmsg = "{}: Boxes value is wrong".format(xlssku)
                    errorlist.append(errmsg)
                
                if str(item['total']).isnumeric() == False:
                    error = True
                    errmsg = "{}: Total value is wrong".format(xlssku)
                    errorlist.append(errmsg)
            if error:
                print("Failed")
                idxdel.append(dlist['name'])
            else:
                print("Passed")
        
        #   DELETE BY idxdel
        for idx in idxdel:
            for index, shipmentdata in enumerate(self.datalist):
                if idx == shipmentdata['name']:
                    del self.datalist[index]
                    break


    @property
    def workbook(self):
        return self.__workbook
   
    @property
    def worksheet(self):
        return self.__worksheet

    @property
    def datalist(self):
        return self.__datalist

    @datalist.setter
    def datalist(self, value):
        self.__datalist = value

    @property
    def datajson(self):
        return json.dumps(self.datalist)  

    @property
    def chrome_data(self):
        return self.__chrome_data
    
    @chrome_data.setter    
    def chrome_data(self, value):
        self.__chrome_data = value

    @property
    def download_folder(self):
        return self.__download_folder
    
    @download_folder.setter    
    def download_folder(self, value):
        self.__download_folder = value

    @property
    def xlsfile(self):
        return self.__xlsfile
    
    @xlsfile.setter    
    def xlsfile(self, value):
        self.__xlsfile = value

    @property
    def file_delimeter(self):
        return self.__delimeter

    @property
    def driver(self):
        return self.__driver

    @property
    def xlworkbook(self):
        return self.__xlworkbook
   
    @property
    def xlworksheet(self):
        return self.__xlworksheet

# def main():
#     parser = argparse.ArgumentParser(description="Amazon Shipment")
#     parser.add_argument('-xls', '--xlsinput', type=str,help="XLSX File Input")
#     parser.add_argument('-sname', '--sheetname', type=str,help="Sheet Name of XLSX file")
#     parser.add_argument('-output', '--pdfoutput', type=str,help="PDF output folder")
#     parser.add_argument('-cdata', '--chromedata', type=str,help="Chrome User Data Directory")
#     args = parser.parse_args()
#     if not (args.xlsinput[-5:] == '.xlsx' or args.xlsinput[-5:] == '.xlsm'):
#         input('2nd File input have to XLSX or XLSM file')
#         sys.exit()
#     isExist = os.path.exists(args.xlsinput)
#     if not isExist:
#         input(args.xlsinput + " does not exist")
#         sys.exit()
#     isExist = os.path.exists(args.chromedata)
#     if isExist == False :
#         input('Please check Chrome User Data Directory')
#         sys.exit()
#     isExist = os.path.exists(args.pdfoutput)
#     if not isExist:
#         input(args.pdfoutput + " folder does not exist")
#         sys.exit()

#     # the second handler is a file handler
#     file_handler = logging.FileHandler('logs/amazonship-err.log')
#     file_handler.setLevel(logging.ERROR)
#     file_handler_format = '%(asctime)s | %(levelname)s | %(lineno)d: %(message)s'
#     file_handler.setFormatter(logging.Formatter(file_handler_format))
#     logger.addHandler(file_handler)

#     file_handler2 = logging.FileHandler('logs/amazonship-info.log')
#     file_handler2.setLevel(logging.INFO)
#     # file_handler2_format = '%(asctime)s | %(levelname)s: %(message)s'
#     file_handler2_format = '%(asctime)s | %(levelname)s | %(lineno)d: %(message)s'
#     file_handler2.setFormatter(logging.Formatter(file_handler2_format))
#     logger2.addHandler(file_handler2)

#     logger2.info("###### Start ######")
#     logger2.info("Filename: {}\nSheet Name:{}\nPDF Output Folder:{}".format(args.xlsinput, args.sheetname, args.pdfoutput))
#     maxrun = 10
#     for i in range(1, maxrun+1):
#         if i > 1:
#             print("Process will be reapeated")
#         try:    
#             shipment = AmazonShipment(xlsfile=args.xlsinput, sname=args.sheetname, chrome_data=args.chromedata, download_folder=args.pdfoutput)
#             shipment.data_sanitizer()
#             if len(shipment.datalist) == 0:
#                 break
#             shipment.parse()
#         except Exception as e:
#             logger.error(e)
#             print("There is an error, check logs/amazonship-err.log")
#             shipment.workbook.save(shipment.xlsfile)
#             shipment.workbook.close()
#             if i == maxrun:
#                 logger.error("Execution Limit reached, Please check the script")
#             continue
#         break
#     addressfile = Path("address.csv")
#     resultfile = lib.join_pdfs(source_folder=args.pdfoutput + lib.file_delimeter() + "combined" , output_folder = args.pdfoutput, tag='Labels')
#     if resultfile != "":
#         lib.add_page_numbers(resultfile)
#         lib.generate_xls_from_pdf(resultfile, addressfile)
#     input("End Process..")    

def main():
    clear_screan()
    parser = argparse.ArgumentParser(description="Amazon Shipment")
    parser.add_argument('-xls', '--xlsinput', type=str,help="XLSX File Input")
    parser.add_argument('-sname', '--sheetname', type=str,help="Sheet Name of XLSX file")
    parser.add_argument('-output', '--pdfoutput', type=str,help="PDF output folder")
    parser.add_argument('-cdata', '--chromedata', type=str,help="Chrome User Data Directory")
    args = parser.parse_args()
    if not (args.xlsinput[-5:] == '.xlsx' or args.xlsinput[-5:] == '.xlsm'):
        input('2nd File input have to XLSX or XLSM file')
        sys.exit()
    isExist = os.path.exists(args.xlsinput)
    if not isExist:
        input(args.xlsinput + " does not exist")
        sys.exit()
    isExist = os.path.exists(args.chromedata)
    if isExist == False :
        input('Please check Chrome User Data Directory')
        sys.exit()
    isExist = os.path.exists(args.pdfoutput)
    if not isExist:
        input(args.pdfoutput + " folder does not exist")
        sys.exit()
    strdate = str(date.today())
    folderamazonship = "{}{}_{}".format(args.pdfoutput + lib.file_delimeter(), 'shipment_creation', strdate) 
    isExist = os.path.exists(folderamazonship)
    if not isExist:
        os.makedirs(folderamazonship)
    
    print('Creating Excel Backup File...', end="", flush=True)
    fnameinput = os.path.basename(args.xlsinput)
    pathinput = args.xlsinput[0:-len(fnameinput)]
    backfile = "{}{}_backup{}".format(pathinput, os.path.splitext(fnameinput)[0], os.path.splitext(fnameinput)[1])
    shutil.copy(args.xlsinput, backfile)
    xltmp = 'xlstmp' + args.xlsinput[-5:]
    try:
        os.remove(xltmp)
    except:
        pass
    shutil.copy(args.xlsinput, xltmp)            

    print('OK')
    print('Opening the Source Excel File...', end="", flush=True)
    xlbook = xw.Book(args.xlsinput)
    print('OK')
    # the second handler is a file handler
    file_handler = logging.FileHandler('logs/amazonship-err.log')
    file_handler.setLevel(logging.ERROR)
    file_handler_format = '%(asctime)s | %(levelname)s | %(lineno)d: %(message)s'
    file_handler.setFormatter(logging.Formatter(file_handler_format))
    logger.addHandler(file_handler)

    file_handler2 = logging.FileHandler('logs/amazonship-info.log')
    file_handler2.setLevel(logging.INFO)
    # file_handler2_format = '%(asctime)s | %(levelname)s: %(message)s'
    file_handler2_format = '%(asctime)s | %(levelname)s | %(lineno)d: %(message)s'
    file_handler2.setFormatter(logging.Formatter(file_handler2_format))
    logger2.addHandler(file_handler2)

    logger2.info("###### Start ######")
    logger2.info("Filename: {}\nSheet Name:{}\nPDF Output Folder:{}".format(args.xlsinput, args.sheetname, folderamazonship))
    maxrun = 10
    for i in range(1, maxrun+1):
        if i > 1:
            print("Process will be reapeated")
        try:    
            shipment = AmazonShipment(xlsfile=args.xlsinput, sname=args.sheetname, chrome_data=args.chromedata, download_folder=folderamazonship, xlworkbook=xlbook)

            shipment.data_sanitizer()
            if len(shipment.datalist) == 0:
                break
            shipment.parse()
            shipment.xlworkbook.save(shipment.xlsfile)
            shipment.workbook.close()
        except Exception as e:
            logger.error(e)
            print("There is an error, check logs/amazonship-err.log")
            # shipment.workbook.save(shipment.xlsfile)
            shipment.xlworkbook.save(shipment.xlsfile)
            shipment.workbook.close()
            if i == maxrun:
                logger.error("Execution Limit reached, Please check the script")
            continue
        break
    addressfile = Path("address.csv")
    resultfile = lib.join_pdfs(source_folder=folderamazonship + lib.file_delimeter() + "combined" , output_folder = folderamazonship, tag='Labels')
    if resultfile != "":
        lib.add_page_numbers(resultfile)
        lib.generate_xls_from_pdf(resultfile, addressfile)
    lib.copysheet(destination=args.xlsinput, source=resultfile[:-4] + ".xlsx", cols=('A', 'B', 'C', 'D', 'E', 'F', 'G', 'H'), sheetsource="Sheet", sheetdestination="Shipment labels summary", tracksheet="dyk_manifest_template", xlbook=xlbook)
    xlbook.save(args.xlsinput)

    input("End Process..")    


if __name__ == '__main__':
    main()
