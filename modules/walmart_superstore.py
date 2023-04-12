from selenium import webdriver
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from webdriver_manager.chrome import ChromeDriverManager as CM
from selenium.webdriver.common.action_chains import ActionChains
import json
import warnings
from openpyxl import Workbook, load_workbook
from urllib.parse import urlencode, urlparse
import time
from random import randint
import pyautogui as pg
import undetected_chromedriver as uc 
import os
import shutil
def getConfig():
	file = open("setting.json", "r")
	config = json.load(file)
	return config


def browser_init():
    config = getConfig()
    warnings.filterwarnings("ignore", category=UserWarning)
    options = webdriver.ChromeOptions()
    # options = Options()
    # options.add_argument("--headless")
    # options.add_argument("user-data-dir={}".format(config['chrome_user_data'])) 
    # options.add_argument("profile-directory={}".format(config['chrome_profile']))
    
    options.add_argument("user-data-dir={}".format("C:\\Users\\User\\AppData\\Local\\Google\\Chrome\\User Data2")) 
    options.add_argument("profile-directory={}".format("Default"))

    options.add_argument('--no-sandbox')
    options.add_argument("--log-level=3")
    # options.add_argument("--window-size=1200, 900")
    options.add_argument('--start-maximized')
    options.add_argument("--disable-notifications")
    options.add_experimental_option("excludeSwitches", ["enable-automation"])
    options.add_experimental_option('useAutomationExtension', False)
    options.add_argument("--disable-blink-features=AutomationControlled")
    # options.add_experimental_option( "prefs",{'profile.managed_default_content_settings.javascript': 1})
    driver = webdriver.Chrome(service=Service(CM().install()), options=options)
    
    return driver

driver = browser_init()
driver.execute_script("Object.defineProperty(navigator, 'webdriver', {get: () => undefined})") 
driver.execute_cdp_cmd("Network.setCacheDisabled", {"cacheDisabled":True})
 
workbook = load_workbook(filename=r"C:/synergy-data-tester/Lookup Listing.xlsx", read_only=False, keep_vba=True, data_only=True)
# workbook = load_workbook(filename="/home/farid/dev/python/synergy-github/data/lookup/Lookup Listing.xlsx", read_only=False, keep_vba=True, data_only=True)
# worksheet = workbook[self.sheetname]
worksheet = workbook["Sheet1"]
user_data = r"C:/Users/User/AppData/Local/Google/Chrome/User Data2"
urlList = []
for i in range(2, worksheet.max_row + 1):
    url = worksheet[f'A{i}'].value
    domain = urlparse(url).netloc

    if domain == 'www.walmart.com' or domain == 'www.walmart.ca':
        urlList.append(url)

i = 0
maxrec = len(urlList)
while True:
    if i == maxrec:
        break
    url = urlList[i]
    print(url, end=" ", flush=True)
    driver.get(url)
    try:
        driver.find_element(By.CSS_SELECTOR, "div#topmessage").text
        print("Failed")
        del driver
        waiting = 120
        print(f'Waiting for bot detection for {waiting} seconds', end=" ", flush=True)
        time.sleep(waiting)
        isExist = os.path.exists(user_data)
        print(isExist)
        if isExist:
            shutil.rmtree(user_data)
        print('OK')
        driver = browser_init()
        continue
    except:
        i += 1
        print('OK')
        pass

    try:
        title = driver.find_element(By.CSS_SELECTOR, "h1[data-automation='product-title']").text
    except:
            title = ''
    try:
        price = driver.find_element(By.CSS_SELECTOR, "span[data-automation='buybox-price']").text
    except:
        price = ''
    try:
        sale = driver.find_element(By.CSS_SELECTOR, "div[data-automation='mix-match-badge'] span").text
    except:
        sale = ''
    print(title, price, sale) 
        
         
